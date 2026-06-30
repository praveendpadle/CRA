import os
import time
import pandas as pd

from curl_cffi import requests
from openpyxl import load_workbook


INPUT_FILE = "input.xlsx"
OUTPUT_FILE = "kroger_output.xlsx"

BASE_URL = "https://www.kroger.com/atlas/v1/stores/v2/locator"

# ---------------------------------
# Session
# ---------------------------------
session = requests.Session(
    impersonate="chrome"
)

session.headers.update({
    "accept": "application/json",
    "referer": "https://www.kroger.com/stores/search",
    "origin": "https://www.kroger.com"
})

# ---------------------------------
# Cache
# ---------------------------------
store_cache = {}


# ---------------------------------
# Load Existing Store IDs
# ---------------------------------
def load_existing_ids():

    if not os.path.exists(OUTPUT_FILE):
        return

    try:

        df = pd.read_excel(
            OUTPUT_FILE,
            usecols=["locationId"]
        )

        for sid in (
            df["locationId"]
            .dropna()
            .astype(str)
        ):

            store_cache[sid] = True

        print(
            f"Loaded {len(store_cache)} existing stores"
        )

    except Exception as ex:

        print("Cache Load Error:", ex)


# ---------------------------------
# Save Row
# ---------------------------------
def save_row(record):

    store_id = str(
        record["locationId"]
    )

    if store_id in store_cache:

        print(f"Skip {store_id}")

        return

    store_cache[store_id] = True

    df = pd.DataFrame([record])

    try:

        if not os.path.exists(
            OUTPUT_FILE
        ):

            df.to_excel(
                OUTPUT_FILE,
                index=False
            )

        else:

            wb = load_workbook(
                OUTPUT_FILE
            )

            ws = wb.active

            with pd.ExcelWriter(
                OUTPUT_FILE,
                mode="a",
                engine="openpyxl",
                if_sheet_exists="overlay"
            ) as writer:

                df.to_excel(
                    writer,
                    index=False,
                    header=False,
                    startrow=ws.max_row
                )

    except Exception as ex:

        if store_id in store_cache:
            del store_cache[store_id]

        print("Save Error:", ex)


# ---------------------------------
# Extract Store
# ---------------------------------
def extract_store(zipcode, s):

    return {

        "InputPincode":
            zipcode,

        "locationId":
            s.get("locationId"),

        "legalName":
            s.get("legalName"),

        "vanityName":
            s.get("vanityName"),

        "storeType":
            s.get("storeType"),

        "managementDivisionNumber":
            s.get("managementDivisionNumber"),

        "address":
            ", ".join(
                s.get("locale", {})
                 .get("address", {})
                 .get("addressLines", [])
            ),

        "city":
            s.get("locale", {})
             .get("address", {})
             .get("cityTown"),

        "state":
            s.get("locale", {})
             .get("address", {})
             .get("stateProvince"),

        "postalCode":
            s.get("locale", {})
             .get("address", {})
             .get("postalCode"),

        "country":
            s.get("locale", {})
             .get("address", {})
             .get("countryCode"),

        "latitude":
            s.get("locale", {})
             .get("location", {})
             .get("lat"),

        "longitude":
            s.get("locale", {})
             .get("location", {})
             .get("lng"),

        "phone":
            s.get("phoneNumber", {})
             .get("pretty"),

        "isOpen":
            s.get("isOpen"),

        "openText":
            s.get("openText"),

        "hours":
            " | ".join(
                f"{h.get('displayName')}: {h.get('displayHours')}"
                for h in s.get(
                    "prettyHours",
                    []
                )
            ),

        "store_url":
            f"https://www.kroger.com/stores/details/{s.get("managementDivisionNumber")}/{s.get('locationId')[3:]}"
    }


# ---------------------------------
# Fetch Stores
# ---------------------------------
def fetch(zipcode):

    response = session.get(

        BASE_URL,

        params={
            "filter.query": zipcode,
            "projections": "full"
        },

        timeout=60
    )

    response.raise_for_status()

    return response.json()


# ---------------------------------
# Main
# ---------------------------------
def process():

    input_df = pd.read_excel(
        INPUT_FILE
    )

    for _, row in input_df.iterrows():

        zipcode = str(
            row["Pincode"]
        ).strip()

       

        print(
            f"\nProcessing {zipcode}"
        )

        try:

            data = fetch(
                zipcode
            )

            stores = (
                data.get(
                    "data",
                    {}
                ).get(
                    "stores",
                    []
                )
            )

            if not stores:

                print(
                    "No stores found"
                )

                continue

            for store in stores:

                record = extract_store(
                    zipcode,
                    store
                )

                save_row(
                    record
                )

                print(
                    f"Saved -> {record['legalName']}"
                )

            time.sleep(3)

        except Exception as ex:

            print(
                f"Error {zipcode}: {ex}"
            )


# ---------------------------------
# Run
# ---------------------------------
load_existing_ids()

process()

print("\nCompleted")