import pandas as pd
import json
import time
import os
import random

from curl_cffi import requests
from openpyxl import load_workbook


INPUT_FILE = "input.xlsx"
OUTPUT_FILE = "harborfreight_output.xlsx"

URL = "https://api.harborfreight.com/graphql"

store_cache = {}


# ===================================
# SESSION
# ===================================
def create_session():

    s = requests.Session(
        impersonate="chrome136"
    )

    s.headers.update({
        "accept": "*/*",
        "origin":
            "https://www.harborfreight.com",
        "referer":
            "https://www.harborfreight.com/store-locator"
    })

    try:
        s.get(
            "https://www.harborfreight.com/store-locator",
            timeout=60
        )
    except:
        pass

    return s


session = create_session()


# ===================================
# RETRY REQUEST
# ===================================
def fetch_with_retry(params):

    global session

    retries = 5

    for attempt in range(retries):

        try:

            response = session.get(
                URL,
                params=params,
                timeout=90
            )

            if response.status_code == 403:

                wait = (
                    10 *
                    (attempt + 1)
                )

                print(
                    f"403 detected "
                    f"Retry {attempt+1}"
                )

                time.sleep(wait)

                try:
                    session.close()
                except:
                    pass

                session = create_session()

                continue

            response.raise_for_status()

            return response

        except Exception as ex:

            wait = (
                5 *
                (attempt + 1)
            )

            print(
                f"Retry Error: {ex}"
            )

            time.sleep(wait)

    return None


# ===================================
# CACHE LOAD
# ===================================
def load_existing_ids():

    if not os.path.exists(
        OUTPUT_FILE
    ):
        return

    try:

        existing = pd.read_excel(
            OUTPUT_FILE,
            usecols=[
                "store_number"
            ]
        )

        for x in (
            existing[
                "store_number"
            ]
            .dropna()
            .astype(str)
        ):

            store_cache[
                x
            ] = True

        print(
            f"Loaded "
            f"{len(store_cache)} "
            f"cached stores"
        )

    except Exception as ex:

        print(
            ex
        )


# ===================================
# SAVE
# ===================================
def append_to_excel(record):

    store_id = str(
        record[
            "store_number"
        ]
    )

    if (
        store_id
        in store_cache
    ):
        return False

    store_cache[
        store_id
    ] = True

    try:

        df = pd.DataFrame(
            [record]
        )

        if not os.path.exists(
            OUTPUT_FILE
        ):

            df.to_excel(
                OUTPUT_FILE,
                index=False
            )

        else:

            wb = load_workbook(
                OUTPUT_FILE
            )

            ws = wb.active

            with pd.ExcelWriter(
                OUTPUT_FILE,
                mode="a",
                engine="openpyxl",
                if_sheet_exists="overlay"
            ) as writer:

                df.to_excel(
                    writer,
                    startrow=ws.max_row,
                    header=False,
                    index=False
                )

        return True

    except Exception:

        if (
            store_id
            in store_cache
        ):
            del store_cache[
                store_id
            ]

        return False


load_existing_ids()

df = pd.read_excel(
    INPUT_FILE
)


# ===================================
# DRIVER
# ===================================
for idx, row in df.iterrows():

    if idx > 0 and idx % 20 == 0:

        print(
            "Refreshing session"
        )

        try:
            session.close()
        except:
            pass

        session = create_session()

    pincode = str(
        row[
            "Pincode"
        ]
    ).strip()

    pincode = pincode.zfill(5)

    print(
        f"\nProcessing {pincode}"
    )

    params = {

        "operationName":
            "FindStoresByQuery",

        "variables":
            json.dumps({

                "filter": {
                    "status":
                        "OPEN"
                },

                "query":
                    pincode,

                "withDistance":
                    True

            }, separators=(",", ":")),

        "extensions":
            json.dumps({

                "persistedQuery": {

                    "version":
                        1,

                    "sha256Hash":
                        (
                            "e61f993d029050546532"
                            "b18a028d529023eeabef"
                            "92e9536a4300c36b516b7f4c"
                        )
                }

            }, separators=(",", ":"))
    }

    response = fetch_with_retry(
        params
    )

    if response is None:

        print(
            f"Skipped {pincode}"
        )

        continue

    try:

        data = response.json()

        root = (
            data
            .get(
                "data",
                {}
            )
            .get(
                "findStoresByQuery",
                {}
            )
        )

        stores = root.get(
            "stores",
            []
        )

        for s in stores:

            saved = append_to_excel({

                "InputPincode":
                    pincode,

                "store_number":
                    s.get(
                        "store_number"
                    ),

                "title":
                    s.get(
                        "title"
                    ),

                "address":
                    s.get(
                        "address"
                    ),

                "telephone":
                    s.get(
                        "telephone"
                    ),

                "latitude":
                    s.get(
                        "latitude"
                    ),

                "longitude":
                    s.get(
                        "longitude"
                    ),

                "distance":
                    s.get(
                        "distance"
                    ),

                "store_url":
                    (
                        "https://www.harborfreight.com/"
                        "storelocator/store"
                        f"?number="
                        f"{s.get('store_number')}"
                    )
            })

            if saved:

                print(
                    f"Saved "
                    f"{s.get('store_number')}"
                )

        time.sleep(
            random.randint(
                5,
                15
            )
        )

    except Exception as ex:

        print(
            ex
        )


print("\nCompleted")