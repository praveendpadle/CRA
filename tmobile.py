import pandas as pd
import time
from curl_cffi import requests
import os
from openpyxl import load_workbook

INPUT_FILE = "input.xlsx"
OUTPUT_FILE = "tmobile_output.xlsx"

BASE_URL = "https://www.t-mobile.com/stores/api/get-nearby-business"

session = requests.Session(
    impersonate="chrome"
)


# =====================================
# SAVE / APPEND
# =====================================
def save_row(record):

    df = pd.DataFrame([record])

    try:

        if not os.path.exists(OUTPUT_FILE):

            df.to_excel(
                OUTPUT_FILE,
                index=False
            )

        else:

            wb = load_workbook(
                OUTPUT_FILE
            )

            ws = wb.active

            with pd.ExcelWriter(
                OUTPUT_FILE,
                mode="a",
                engine="openpyxl",
                if_sheet_exists="overlay"
            ) as writer:

                df.to_excel(
                    writer,
                    index=False,
                    header=False,
                    startrow=ws.max_row
                )

    except Exception as ex:

        print(
            "Save Error:",
            ex
        )


# =====================================
# DRIVER
# =====================================

df = pd.read_excel(
    INPUT_FILE
)

for _, row in df.iterrows():

    pincode = str(
        row["Pincode"]
    ).strip()

    state = str(
        row["State"]
    ).strip()

    country = str(
        row["Country"]
    ).strip()

    lat = row["Lat"]
    lng = row["Lng"]

    page = 1

    loc = f"{state} {pincode}, {country}"

    print(f"\nProcessing {loc}")

    while True:

        try:

            response = session.get(

                BASE_URL,

                params={
                    "INTNAV":
                        "tNav:StoreLocator",

                    "page":
                        page,

                    "loc":
                        loc,

                    "lat":
                        lat,

                    "lon":
                        lng,

                    "bsort":
                        "recommended"
                },

                headers={
                    "Referer":
                        "https://www.t-mobile.com/stores",

                    "Accept":
                        "application/json"
                },

                timeout=60
            )

            print(
                "Status:",
                response.status_code
            )

            response.raise_for_status()

            data = response.json()

            business = data.get(
                "business_list",
                {}
            )

            stores = business.get(
                "object_list",
                []
            )

            if not stores:
                break

            for item in stores:

                record = {

                    "InputPincode":
                        pincode,

                    "display_name":
                        item.get(
                            "display_name"
                        ),

                    "external_store_code":
                        item.get(
                            "external_store_code"
                        ),

                    "lat":
                        item.get(
                            "lat"
                        ),

                    "lon":
                        item.get(
                            "lon"
                        ),

                    "address_postcode":
                        item.get(
                            "address_postcode"
                        ),

                    "schemaHrs":
                        ", ".join(
                            item.get(
                                "all_opening_hours",
                                {}
                            ).get(
                                "schemaHrs",
                                []
                            )
                        ),

                    "product_link":
                       f"https://www.t-mobile.com{item.get('product_link')}",

                    "business_phone_text":
                        item.get(
                            "contact_context",
                            {}
                        ).get(
                            "business_phone_text"
                        ),

                    "address_text":
                        item.get(
                            "address_text"
                        ),

                    "get_directions_link":
                        item.get(
                            "get_directions_link"
                        )
                }

                # SAVE IMMEDIATELY
                save_row(
                    record
                )

                print(
                    f"Saved -> {record['display_name']}"
                )

            next_page = business.get(
                "next_page_number"
            )

            print(
                "Next Page:",
                next_page
            )

            if next_page is None:
                break

            page = next_page

            time.sleep(1)

        except Exception as ex:

            print(
                f"Failed: {ex}"
            )

            break


print("\nDone")