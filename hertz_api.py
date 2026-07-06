import os
import time
import pandas as pd
from openpyxl import Workbook, load_workbook
from curl_cffi import requests

INPUT_FILE = "input.xlsx"
OUTPUT_FILE = "output_hertz_api.xlsx"

URL = "https://ecom.mss.hertz.io/mdm-locations/internal-lookup/geo-search/hertz/"


# -------------------------------
# Session
# -------------------------------
session = requests.Session(
    impersonate="chrome136"
)

session.headers.update({
    "accept": "application/json",
    "user-agent": "Mozilla/5.0",
    "origin": "https://www.hertz.com",
    "referer": "https://www.hertz.com/"
})


# -------------------------------
# Create output workbook
# -------------------------------
if not os.path.exists(OUTPUT_FILE):

    wb = Workbook()
    ws = wb.active

    ws.append([
        "SearchZip",
        "OAG",
        "OAG3",
        "WWD Code",
        "Name",
        "Location Type",
        "Ownership",
        "Operations Type",
        "Distance",

        "Address1",
        "Address2",
        "City",
        "State",
        "PostalCode",
        "Country",
        "Latitude",
        "Longitude",
        "Full Address",

        "Phone",
        "International Phone",
        "Email",
        "Fax",

        "Hours",

        "Website",

        "IsOpen",
        "BookingLocation",
        "Airport",
        "WalkIns",
        "24x7",

        "Raw JSON"
    ])

    wb.save(OUTPUT_FILE)


# -------------------------------
# Load workbook
# -------------------------------
wb = load_workbook(OUTPUT_FILE)
ws = wb.active


# -------------------------------
# Existing IDs
# -------------------------------
existing = set()

for row in ws.iter_rows(min_row=2, values_only=True):

    if row[1]:
        existing.add(str(row[1]))


print("Existing Locations:", len(existing))


# -------------------------------
# Read input ZIPs
# -------------------------------
df = pd.read_excel(INPUT_FILE, dtype={0: str})

zip_column = df.columns[0]

zipcodes = (
    df[zip_column]
    .dropna()
    .astype(str)
    .str.strip()
    .str.zfill(5) 
    .tolist()
)

print("ZIP Codes:", len(zipcodes))


# -------------------------------
# Process ZIPs
# -------------------------------
for zipcode in zipcodes:

    print(f"\nSearching {zipcode}")

    success = False

    for retry in range(3):

        try:

            response = session.get(
                URL,
                params={
                    "search": zipcode,
                    "radius": 150
                },
                timeout=60
            )

            if response.status_code != 200:

                print("Status:", response.status_code)
                time.sleep(2)
                continue

            data = response.json()

            success = True
            break

        except Exception as ex:

            print(ex)
            time.sleep(2)

    if not success:
        print("Failed:", zipcode)
        continue

    locations = data.get("data", [])

    print("Locations:", len(locations))

    for loc in locations:

        oag = loc.get("oag")

        if not oag:
            continue

        if oag in existing:
            continue

        existing.add(oag)

        address = loc.get("address", {})
        contact = loc.get("contact_info", {})
        phone = contact.get("phone", {})
        website = loc.get("website_urls", {})
        open_status = loc.get("open_status", {})

        ws.append([

            zipcode,

            oag,
            loc.get("oag3"),
            loc.get("wwd_legacy_code"),

            loc.get("name"),
            loc.get("location_type"),
            loc.get("ownership_type"),
            loc.get("operations_type"),
            loc.get("distance"),

            address.get("address1"),
            address.get("address2"),
            address.get("city"),
            address.get("state_short"),
            address.get("postal_code"),
            address.get("country_short"),
            address.get("latitude"),
            address.get("longitude"),
            address.get("full_address"),

            phone.get("national_phone_number"),
            phone.get("international_phone_number"),
            contact.get("email"),
            contact.get("fax"),

            loc.get("hours_of_operation_1"),

            website.get("full_url"),

            open_status.get("is_open"),
            loc.get("is_bookinglocation"),
            loc.get("is_onairport"),
            loc.get("allow_walk_ins"),
            loc.get("is_24_7"),

            str(loc)

        ])

    wb.save(OUTPUT_FILE)

    print("Saved")

    time.sleep(1)

wb.save(OUTPUT_FILE)

print("\nCompleted")
print("Unique Locations:", len(existing))