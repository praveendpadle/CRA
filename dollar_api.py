import os
import time
import pandas as pd
from openpyxl import Workbook, load_workbook
from curl_cffi import requests

INPUT_FILE = "input.xlsx"
OUTPUT_FILE = "output.xlsx"

URL = "https://ecom.mss.hertz.io/mdm-locations/internal-lookup/geo-search/dollar/"

# -------------------------------------------------------
# Day Mapping
# -------------------------------------------------------

DAY_MAP = {
    "1": "Monday",
    "2": "Tuesday",
    "3": "Wednesday",
    "4": "Thursday",
    "5": "Friday",
    "6": "Saturday",
    "7": "Sunday"
}


def format_hours(curated_days):
    """
    Convert curated_days to readable format.
    """

    if not curated_days:
        return ""

    output = []

    for day in range(1, 8):

        key = str(day)

        day_name = DAY_MAP[key]

        periods = curated_days.get(key, [])

        if not periods:
            output.append(f"{day_name}: Closed")
            continue

        ranges = []

        for period in periods:

            start = period.get("start_time", "")
            end = period.get("end_time", "")

            if start and end:
                ranges.append(f"{start}-{end}")

        if ranges:
            output.append(f"{day_name}: {', '.join(ranges)}")
        else:
            output.append(f"{day_name}: Closed")

    return " | ".join(output)


# -------------------------------------------------------
# Session
# -------------------------------------------------------

session = requests.Session(
    impersonate="chrome136"
)

session.headers.update({
    "accept": "application/json",
    "user-agent": "Mozilla/5.0",
    "origin": "https://www.dollar.com",
    "referer": "https://www.dollar.com/"
})

# -------------------------------------------------------
# Create Output
# -------------------------------------------------------

if not os.path.exists(OUTPUT_FILE):

    wb = Workbook()
    ws = wb.active

    ws.append([
        "SearchZip",
        "OAG",
        "Name",
        "LocationType",
        "Ownership",
        "OperationsType",
        "Distance",

        "Address1",
        "Address2",
        "City",
        "State",
        "PostalCode",
        "Country",

        "Latitude",
        "Longitude",

        "Phone",
        "Email",

        "OperatingHours",

        "Website",

        "Airport",
        "BookingLocation",

        "RawJSON"
    ])

    wb.save(OUTPUT_FILE)

# -------------------------------------------------------
# Load Workbook
# -------------------------------------------------------

wb = load_workbook(OUTPUT_FILE)
ws = wb.active

existing = set()

for row in ws.iter_rows(min_row=2, values_only=True):

    if row[1]:
        existing.add(str(row[1]))

print("Existing:", len(existing))

# -------------------------------------------------------
# Read ZIP Codes
# -------------------------------------------------------

df = pd.read_excel(INPUT_FILE, dtype={0: str})

zip_column = df.columns[0]

zipcodes = (
    df[zip_column]
    .fillna("")
    .astype(str)
    .str.strip()
    .str.split(".").str[0]
    .str.zfill(5)
)

zipcodes = [
    z for z in zipcodes
    if z.isdigit() and len(z) == 5
]

print("ZIP Codes:", len(zipcodes))

# -------------------------------------------------------
# Process
# -------------------------------------------------------

for zipcode in zipcodes:

    print(f"\nSearching {zipcode}")

    success = False

    for retry in range(3):

        try:

            response = session.get(
                URL,
                params={
                    "search": zipcode,
                    "radius": 150
                },
                timeout=60
            )

            if response.status_code != 200:

                print("Status:", response.status_code)
                time.sleep(2)
                continue

            data = response.json()

            success = True
            break

        except Exception as ex:

            print(ex)
            time.sleep(2)

    if not success:
        continue

    locations = data.get("data", [])

    print("Locations:", len(locations))

    for loc in locations:

        oag = loc.get("oag")

        if not oag:
            continue

        if oag in existing:
            continue

        existing.add(oag)

        address = loc.get("address", {})
        contact = loc.get("contact_info", {})
        phone = contact.get("phone", {})
        website = loc.get("website_urls", {})

        hours = format_hours(
            loc.get("curated_days", {})
        )

        ws.append([

            zipcode,

            oag,

            loc.get("name"),
            loc.get("location_type"),
            loc.get("ownership_type"),
            loc.get("operations_type"),
            loc.get("distance"),

            address.get("address1"),
            address.get("address2"),
            address.get("city"),
            address.get("state_short"),
            address.get("postal_code"),
            address.get("country_short"),

            address.get("latitude"),
            address.get("longitude"),

            phone.get("national_phone_number"),
            contact.get("email"),

            hours,

            website.get("full_url"),

            loc.get("is_onairport"),
            loc.get("is_bookinglocation"),

            str(loc)

        ])

    wb.save(OUTPUT_FILE)

    print("Saved")

    time.sleep(0.5)

wb.save(OUTPUT_FILE)

print("\nCompleted")
print("Total Unique Locations:", len(existing))