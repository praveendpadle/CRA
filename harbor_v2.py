import pandas as pd
import json
import time
import os

from curl_cffi import requests
from openpyxl import load_workbook


INPUT_FILE = "input.xlsx"
OUTPUT_FILE = "harborfreight_output.xlsx"

URL = "https://api.harborfreight.com/graphql"

session = requests.Session(
    impersonate="chrome136"
)

session.headers.update({
    "accept": "*/*",
    "origin": "https://www.harborfreight.com",
    "referer": "https://www.harborfreight.com/store-locator"
})

# ===================================
# CACHE
# store_number -> True
# ===================================
store_cache = {}


# ===================================
# LOAD EXISTING IDS
# ===================================
def load_existing_ids():

    if not os.path.exists(
        OUTPUT_FILE
    ):
        return

    try:

        existing = pd.read_excel(
            OUTPUT_FILE,
            usecols=[
                "store_number"
            ]
        )

        for x in (
            existing[
                "store_number"
            ]
            .dropna()
            .astype(str)
        ):

            store_cache[
                x
            ] = True

        print(
            f"Loaded "
            f"{len(store_cache)} "
            f"cached stores"
        )

    except Exception as ex:

        print(
            "Cache Load Error:",
            ex
        )


# ===================================
# APPEND
# ===================================
def append_to_excel(record):

    store_id = str(
        record.get(
            "store_number",
            ""
        )
    ).strip()

    # Skip duplicate
    if (
        store_id
        in store_cache
    ):

        print(
            f"Skip {store_id}"
        )

        return False

    # Add to cache
    store_cache[
        store_id
    ] = True

    try:

        df_new = pd.DataFrame(
            [record]
        )

        if not os.path.exists(
            OUTPUT_FILE
        ):

            df_new.to_excel(
                OUTPUT_FILE,
                index=False
            )

        else:

            wb = load_workbook(
                OUTPUT_FILE
            )

            ws = wb.active

            with pd.ExcelWriter(
                OUTPUT_FILE,
                mode="a",
                engine="openpyxl",
                if_sheet_exists="overlay"
            ) as writer:

                df_new.to_excel(
                    writer,
                    index=False,
                    header=False,
                    startrow=ws.max_row
                )

        return True

    except Exception as ex:

        # rollback cache
        if (
            store_id
            in store_cache
        ):
            del store_cache[
                store_id
            ]

        print(
            "Save Error:",
            ex
        )

        return False


# ===================================
# LOAD CACHE
# ===================================
load_existing_ids()

# ===================================
# INPUT
# ===================================
df = pd.read_excel(
    INPUT_FILE
)

for _, row in df.iterrows():

    pincode = str(
        row[
            "Pincode"
        ]
    ).strip()

    print(
        f"\nProcessing {pincode}"
    )

    variables = {

        "filter": {
            "status":
                "OPEN"
        },

        "query":
            pincode,

        "withDistance":
            True
    }

    extensions = {

        "persistedQuery": {

            "version":
                1,

            "sha256Hash":
                (
                    "e61f993d029050546532"
                    "b18a028d529023eeabef"
                    "92e9536a4300c36b516b7f4c"
                )
        }
    }

    try:

        response = session.get(

            URL,

            params={

                "operationName":
                    "FindStoresByQuery",

                "variables":
                    json.dumps(
                        variables,
                        separators=(
                            ",",
                            ":"
                        )
                    ),

                "extensions":
                    json.dumps(
                        extensions,
                        separators=(
                            ",",
                            ":"
                        )
                    )
            },

            timeout=60
        )

        response.raise_for_status()

        data = response.json()

        root = (
            data
            .get(
                "data",
                {}
            )
            .get(
                "findStoresByQuery",
                {}
            )
        )

        search_lat = (
            root.get(
                "latitude"
            )
        )

        search_lon = (
            root.get(
                "longitude"
            )
        )

        stores = (
            root.get(
                "stores",
                []
            )
        )

        for s in stores:

            record = {

                "InputPincode":
                    pincode,

                "search_latitude":
                    search_lat,

                "search_longitude":
                    search_lon,

                "store_number":
                    s.get(
                        "store_number"
                    ),

                "title":
                    s.get(
                        "title"
                    ),

                "address":
                    s.get(
                        "address"
                    ),

                "address_description":
                    s.get(
                        "address_description"
                    ),

                "city":
                    s.get(
                        "city"
                    ),

                "state":
                    s.get(
                        "state"
                    ),

                "postcode":
                    s.get(
                        "postcode"
                    ),

                "latitude":
                    s.get(
                        "latitude"
                    ),

                "longitude":
                    s.get(
                        "longitude"
                    ),

                "telephone":
                    s.get(
                        "telephone"
                    ),

                "status":
                    s.get(
                        "status"
                    ),

                "store_hours_mf":
                    s.get(
                        "store_hours_mf"
                    ),

                "store_hours_sat":
                    s.get(
                        "store_hours_sat"
                    ),

                "store_hours_sun":
                    s.get(
                        "store_hours_sun"
                    ),

                "distance":
                    s.get(
                        "distance"
                    ),

                "soft_opening_date":
                    s.get(
                        "soft_opening_date"
                    ),

                "image":
                    s.get(
                        "image"
                    ),

                "store_url":
                    (
                        "https://www.harborfreight.com/"
                        "storelocator/store"
                        f"?number="
                        f"{s.get('store_number')}"
                    )
            }

            saved = append_to_excel(
                record
            )

            if saved:

                print(
                    f"Saved "
                    f"{record['store_number']}"
                )

        time.sleep(1)

    except Exception as ex:

        print(
            f"Error {pincode}: {ex}"
        )


print("\nCompleted")
