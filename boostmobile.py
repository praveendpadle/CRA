import pandas as pd
import time
import os

from curl_cffi import requests
from openpyxl import load_workbook


INPUT_FILE = "input.xlsx"
OUTPUT_FILE = "boostmobile_output.xlsx"

BASE_URL = (
    "https://www.boostmobile.com/locations/api/get-nearby-business"
)

session = requests.Session(
    impersonate="chrome"
)

# =====================================
# CACHE
# external_store_code -> True
# =====================================
store_cache = {}


# =====================================
# LOAD EXISTING IDS
# =====================================
def load_existing_ids():

    if not os.path.exists(
        OUTPUT_FILE
    ):
        return

    try:

        df = pd.read_excel(
            OUTPUT_FILE,
            usecols=[
                "external_store_code"
            ]
        )

        for x in (
            df[
                "external_store_code"
            ]
            .dropna()
            .astype(str)
        ):

            store_cache[
                x
            ] = True

        print(
            f"Loaded "
            f"{len(store_cache)} "
            f"cached stores"
        )

    except Exception as ex:

        print(
            "Cache Load Error:",
            ex
        )


# =====================================
# SAVE / APPEND
# =====================================
def save_row(record):

    store_id = str(
        record.get(
            "external_store_code",
            ""
        )
    ).strip()

    # Skip duplicates
    if (
        store_id
        in store_cache
    ):

        print(
            f"Skip -> {store_id}"
        )

        return False

    store_cache[
        store_id
    ] = True

    df = pd.DataFrame(
        [record]
    )

    try:

        if not os.path.exists(
            OUTPUT_FILE
        ):

            df.to_excel(
                OUTPUT_FILE,
                index=False
            )

        else:

            wb = load_workbook(
                OUTPUT_FILE
            )

            ws = wb.active

            with pd.ExcelWriter(
                OUTPUT_FILE,
                mode="a",
                engine="openpyxl",
                if_sheet_exists="overlay"
            ) as writer:

                df.to_excel(
                    writer,
                    index=False,
                    header=False,
                    startrow=ws.max_row
                )

        return True

    except Exception as ex:

        print(
            "Save Error:",
            ex
        )

        # rollback cache
        if (
            store_id
            in store_cache
        ):
            del store_cache[
                store_id
            ]

        return False


# =====================================
# LOAD CACHE
# =====================================
load_existing_ids()


# =====================================
# PROCESS
# =====================================

df = pd.read_excel(
    INPUT_FILE
)

for _, row in df.iterrows():

    pincode = str(
        row["Pincode"]
    ).strip()

    state = str(
        row["State"]
    ).strip()

    country = str(
        row["Country"]
    ).strip()

    lat = row["Lat"]
    lng = row["Lng"]

    page = 1

    loc = (
        f"{state} "
        f"{pincode}, "
        f"{country}"
    )

    print(
        f"\nProcessing {loc}"
    )

    while True:

        try:

            response = session.get(

                BASE_URL,

                params={

                    "INTNAV":
                        "tNav:StoreLocator",

                    "page":
                        page,

                    "loc":
                        loc,

                    "lat":
                        lat,

                    "lon":
                        lng,

                    "bsort":
                        "recommended"
                },

                headers={

                    "Referer":
                        "https://www.boostmobile.com/locations",

                    "Accept":
                        "application/json"
                },

                timeout=60
            )

            print(
                "Status:",
                response.status_code
            )

            response.raise_for_status()

            data = response.json()

            business = data.get(
                "business_list",
                {}
            )

            stores = business.get(
                "object_list",
                []
            )

            if not stores:

                print(
                    "No stores"
                )

                break

            for item in stores:

                record = {

    "InputPincode":
        pincode,

    "external_store_code":
        item.get(
            "external_store_code"
        ),

    "display_name":
        item.get(
            "display_name"
        ),

    "lat":
        item.get(
            "lat"
        ),

    "lon":
        item.get(
            "lon"
        ),

    "address_postcode":
        item.get(
            "address_postcode"
        ),

    "schemaHrs":
        ", ".join(
            item
            .get(
                "all_opening_hours",
                {}
            )
            .get(
                "schemaHrs",
                []
            )
        ),

    # NULL IF EMPTY
    "business_link":
        (
            None
            if not item.get(
                "business_link"
            )
            else (
                "https://www.boostmobile.com"
                +
                str(
                    item.get(
                        "business_link"
                    )
                )
            )
        ),

    "business_phone_text":
        (
            item
            .get(
                "contact_context",
                {}
            )
            .get(
                "business_phone_text"
            )
        ),

    "address_text":
        item.get(
            "address_text"
        ),

    "get_directions_link":
        item.get(
            "get_directions_link"
        ),

    # NEW KEYS
    "non_storefront_url":
        item.get(
            "non_storefront_url"
        ),

    "business_type_text":
        item.get(
            "business_type_text"
        ),

    "business_type_url_name":
        item.get(
            "business_type_url_name"
        ),

    "permanently_closed":
        item.get(
            "permanently_closed"
        ),

    "temporarily_closed":
        item.get(
            "temporarily_closed"
        )
}

                saved = save_row(
                    record
                )

                if saved:

                    print(
                        f"Saved -> "
                        f"{record['display_name']}"
                    )

            next_page = business.get(
                "next_page_number"
            )

            print(
                "Next Page:",
                next_page
            )

            if next_page is None:
                break

            page = next_page

            time.sleep(1)

        except Exception as ex:

            print(
                f"Failed: {ex}"
            )

            break


print("\nDone")