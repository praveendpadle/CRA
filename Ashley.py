import time
import pandas as pd

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support import expected_conditions as EC
import os
from openpyxl import load_workbook

BASE = "https://stores.ashleyfurniture.com"
START = BASE + "/store?type=all"


OUTPUT_FILE = "ashley_store_details.xlsx"


#SAVE FILE
def save_row(row):

    df = pd.DataFrame([row])

    try:

        if not os.path.exists(
            OUTPUT_FILE
        ):

            df.to_excel(
                OUTPUT_FILE,
                index=False
            )

        else:

            book = load_workbook(
                OUTPUT_FILE
            )

            sheet = book.active

            start_row = (
                sheet.max_row
            )

            with pd.ExcelWriter(

                OUTPUT_FILE,

                engine="openpyxl",

                mode="a",

                if_sheet_exists="overlay"

            ) as writer:

                writer._book = book

                df.to_excel(

                    writer,

                    startrow=start_row,

                    header=False,

                    index=False

                )

    except Exception as ex:

        print(
            "Save Error:",
            ex
        )
# =====================================
# DRIVER
# =====================================
def create_driver():

    options = Options()

    # options.add_argument("--headless=new")

    options.add_argument("--start-maximized")

    driver = webdriver.Chrome(
        service=Service(
            ChromeDriverManager().install()
        ),
        options=options
    )

    return driver


driver = create_driver()


# =====================================
# OPEN
# =====================================
def load(url):

    print("\nOPEN:", url)

    driver.get(url)

    WebDriverWait(
        driver,
        30
    ).until(
        lambda d:
        d.execute_script(
            "return document.readyState"
        ) == "complete"
    )

    time.sleep(2)


# =====================================
# SAFE HELPERS
# =====================================
def text(css):

    try:
        return driver.find_element(
            By.CSS_SELECTOR,
            css
        ).text.strip()

    except:
        return ""


def attr(css, name):

    try:

        return (
            driver
            .find_element(
                By.CSS_SELECTOR,
                css
            )
            .get_attribute(name)
        )

    except:

        return ""


# =====================================
# STATES
# =====================================
def get_states():

    load(START)

    states = []

    rows = driver.find_elements(
        By.CSS_SELECTOR,
        ".tabbed-content .state-col a"
    )

    for r in rows:

        try:

            url = r.get_attribute(
                "href"
            )

            state = (
                r.find_element(
                    By.CSS_SELECTOR,
                    ".stateName"
                )
                .text
                .split("(")[0]
                .strip()
            )

            states.append({
                "state": state,
                "url": url
            })

        except:
            pass

    return states


# =====================================
# STORE LINKS
# =====================================
def get_store_links(
    state_url
):

    load(state_url)

    urls = []

    links = driver.find_elements(
        By.XPATH,
        "//a[contains(@href,'/store/us/')]"
    )

    for x in links:

        href = x.get_attribute(
            "href"
        )

        if (
            href
            and href != state_url
            and href not in urls
        ):

            urls.append(
                href
            )

    return urls


# =====================================
# HOURS
# =====================================

def get_hours():

    try:
        time.sleep(3)
        # Expand Store Hours accordion
        try:

            accordion = driver.find_element(
                By.CSS_SELECTOR,
                "#cardStoreHours .accordion-title"
            )

            expanded = (
                accordion.get_attribute(
                    "aria-expanded"
                )
                or "false"
            )

            if expanded.lower() == "false":

                driver.execute_script(
                    "arguments[0].click();",
                    accordion
                )
            time.sleep(3)
        except:
            pass


        # Wait until hours are rendered
        WebDriverWait(
            driver,
            10
        ).until(

            lambda d:
            d.find_element(
                By.ID,
                "storeHoursContent"
            ).is_displayed()

        )


        rows = driver.find_elements(
            By.CSS_SELECTOR,
            "#storeHours .day"
        )

        result = []

        for row in rows:

            try:

                day = (
                    row.find_element(
                        By.CSS_SELECTOR,
                        ".text"
                    )
                    .text
                    .strip()
                )

                hours = (
                    row.find_element(
                        By.CSS_SELECTOR,
                        ".hours"
                    )
                    .text
                    .strip()
                )

                result.append(
                    f"{day}: {hours}"
                )

            except:
                continue


        return " | ".join(result)

    except Exception as ex:

        print(
            "Hours Error:",
            ex
        )

        return ""

# =====================================
# STORE DETAILS
# =====================================
def parse_store(
    url,
    state
):

    load(url)

    WebDriverWait(
        driver,
        15
    ).until(

        lambda d:
        d.find_elements(
            By.ID,
            "location-details"
        )

    )

    root = driver.find_element(
        By.ID,
        "location-details"
    )

    store_no = root.get_attribute(
        "data-id"
    )

    lat = root.get_attribute(
        "data-lat"
    )

    lng = root.get_attribute(
        "data-lng"
    )

    # name = text(
    #     ".location-name"
    # )
    try:

        store_name = driver.find_element(
            By.ID,
            "storeName"
        )

        full_name = (
            store_name.text
            .replace("\n", " ")
            .strip()
        )

    except:

        full_name = ""
    phone = attr(
        ".phone.location-page a",
        "data-phone"
    )

    addr1 = text(
        ".address.address1"
    )

    addr2 = text(
        ".address.address2"
    )

    address = (
        f"{addr1}, {addr2}"
        if addr2
        else addr1
    )

    maps = attr(
        "a.directions",
        "href"
    )

    hours = get_hours()

    return {

        "Store No":
            store_no,

        "State":
            state,

        "Name":
            full_name,

        "Latitude":
            lat,

        "Longitude":
            lng,

        "Phone":
            phone,

        "Address":
            address,

        "Operating Hours":
            hours,

        "Google Map Link":
            maps,

        "Store URL":
            url

    }


# =====================================
# MAIN
# =====================================
rows = []

try:

    states = get_states()

    print(
        "States:",
        len(states)
    )

    for st in states:

        print(
            "\nSTATE:",
            st["state"]
        )

        try:

            stores = get_store_links(
                st["url"]
            )

            print(
                "Stores:",
                len(stores)
            )

            for s in stores:

                try:

                    # rows.append(

                    #     parse_store(
                    #         s,
                    #         st["state"]
                    #     )

                    # )
                    row = parse_store(
                    s,
                    st["state"]
                )

                    if row:

                        rows.append(
                            row
                        )

                        save_row(
                            row
                        )

                        print(
                            row
                        )

                except Exception as ex:

                    print(
                        "Store Error:",
                        ex
                    )

        except Exception as ex:

            print(
                "State Error:",
                ex
            )

finally:

    driver.quit()


# =====================================
# EXPORT
# =====================================

print("\nCompleted")
print(
    "Total:",
    len(rows)
)
print(
    "Saved:",
    OUTPUT_FILE
)


# df = pd.DataFrame(
#     rows
# )

# outfile = (
#     "ashley_store_details.xlsx"
# )

# df.to_excel(
#     outfile,
#     index=False
# )

# print("\nDONE")

# print(
#     "Rows:",
#     len(df)
# )

# print(
#     "Saved:",
#     outfile
# )